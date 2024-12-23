import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
import zeyrek
import string
import pandas as pd
from nltk.corpus import stopwords
import openpyxl
from openpyxl.styles import PatternFill
import rules

# Gerekli NLTK paketlerini indirin
#nltk.download("all")
#nltk.download('punkt')
#nltk.download('stopwords')


def load_excel_file(file_path):
    """
    Excel dosyasını yükleyip, 'Cümle' sütununu döndüren fonksiyon.
    """
    df = pd.read_excel(file_path, engine='openpyxl')
    return df['Cümle'].tolist()  # 'Cümle' sütununu listeye dönüştür

def initialize_analyzers():
    """
    Zeyrek ve Stanza analizörlerini başlatan fonksiyon.
    """
    analyzer = zeyrek.MorphAnalyzer()
    return analyzer

def remove_punctuation(text):
    """
    Noktalama işaretlerini kaldıran fonksiyon.
    """
    words = word_tokenize(text)
    words = [word for word in words if word not in string.punctuation]
    return ' '.join(words)

def process_text_with_zeyrek(text, analyzer):
    """
    Zeyrek analizörü ile metni işleyen fonksiyon. Kelime, kök, ekler ve pos değerlerini döndürür.
    """
    sentences = sent_tokenize(text)  # Metni cümlelere ayır
    processed_sentences = []

    # Türkçe StopWords listesini al
    turkish_stopwords = set(stopwords.words('turkish'))

    for sentence in sentences:
        words = word_tokenize(sentence)  # Kelimelere ayır

        # Zeyrek ile kelimeleri köklere, eklerine ve POS değerine ayır
        sentence_info = []
        for word in words:
            word_lower = word.lower()  # Büyük-küçük harf duyarlılığını kaldır
            if word_lower in turkish_stopwords:  # Stopwords listesinde olan kelimeleri atla
                continue
            analysis = analyzer.analyze(word)
            if analysis:
                root = analysis[0][0].lemma  # Kelimenin kökü
                suffixes = "+".join(analysis[0][0].morphemes[1:])  # Kelimenin ekleri
                pos = analysis[0][0].pos  # Kelimenin türü (POS)
                sentence_info.append({
                    'Kelime': word,
                    'Kök': root,
                    'Ekler': suffixes if suffixes else 'Yok',
                    'POS': pos
                })
            else:
                sentence_info.append({
                    'Kelime': word,
                    'Kök': word,
                    'Ekler': 'Yok',
                    'POS': 'Bilinmiyor'  # Eğer POS bulunamazsa
                })

        processed_sentences.append(sentence_info)

    return processed_sentences

def display_processed_text(processed_text):
    """
    İşlenmiş metni formatlı şekilde yazdıran fonksiyon.
    """
    print("----------------------------------------------------------------------------------------------------------")
    for i, sentence in enumerate(processed_text):
        print(f"Cümle {i+1}:")
        for word_info in sentence:
            print(f"Kelime: {word_info['Kelime']}, \nKök: {word_info['Kök']}, \nEkler: {word_info['Ekler']}, \nPOS: {word_info['POS']}\n--------------------------------------")
        print("----------------------------------------------------------------------------------------------------------")


def analyze_sentence(processed_text):
    # Flagleri başlat
    input_data = {
        "negation_suffix": False,
        "double_negation": False,
        "positive_words": False,
        "negative_words": False,
        "positive_score_gte_negative": False,
        "negative_score_gt_positive": False,
        "positive_degil" : False,
        "negative_degil" : False,
        "ironic_punctuation": False,
        "negative_beforeComma" : False,
        "positive_beforeComma" : False,

        "no_negation_suffix": True,
        "no_double_negation": True,
        "no_positive_words": True,
        "no_negative_words":True,
        "no_positive_degil":True,
        "no_negative_degil":True,
        "no_ironic_punctuation":True,
        "no_negative_beforeComma" : True,
        "no_positive_beforeComma" : True,
    }

    input_data["negation_suffix"] = rules.negation_suffix_control(processed_text)
    input_data["double_negation"] = rules.double_negation_control(processed_text,"polarity_negative.txt")
    
    input_data["positive_degil"] = rules.positive_degil_control(processed_text, "polarity_positive.txt")
    input_data["negative_degil"] = rules.negative_degil_control(processed_text, "polarity_negative.txt")

    input_data["positive_words"] = rules.positive_polarity_control(processed_text, "polarity_positive.txt")
    input_data["negative_words"] = rules.negative_polarity_control(processed_text, "polarity_negative.txt")

    input_data["negative_beforeComma"] = rules.check_before_comma(processed_text,"polarity_negative.txt")
    input_data["positive_beforeComma"] = rules.check_before_comma(processed_text,"polarity_positive.txt")

    input_data["ironic_punctuation"] = rules.ironic_punctuation(processed_text)

    positive_score = rules.positive_score_calculate(processed_text, "polarity_positive.txt")
    negative_score = rules.negative_score_calculate(processed_text, "polarity_negative.txt")

    print("Positive Polarity : " , rules.positive_polarity_control(processed_text, "polarity_positive.txt") , "\t Score : ", positive_score)
    print("Negative Polarity : " , rules.negative_polarity_control(processed_text, "polarity_negative.txt") , "\t Score : ", negative_score)

    if(positive_score >= negative_score):
        input_data["positive_score_gte_negative"] = True
    
    if(negative_score > positive_score):
        input_data["negative_score_gt_positive"] = True

    return input_data

def run_fsm(fsm, input_data):
    """
    FSM'i çalıştırır ve bir son duruma ("positive" veya "negative") ulaşana kadar devam eder.
    """
    current_state = "start"
    print(f"Başlangıç Durumu: {current_state}")

    while current_state not in ["Pozitif", "Negatif"]:
        print(f"Mevcut Durum: {current_state}")

        # Mevcut durumdaki geçişleri al
        transitions = fsm.get(current_state, {})
        transition_found = False

        # Geçiş koşullarını kontrol et ve uygun bir sonraki duruma ilerle
        for condition, next_state in transitions.items():
            if condition in input_data and input_data[condition]:
                print(f"Koşul Sağlandı: {condition} -> Geçiş Yapılıyor: {next_state}")
                current_state = next_state
                transition_found = True
                break
        
        if not transition_found:
            print(f"Durum '{current_state}' için geçerli bir geçiş bulunamadı.")
            raise ValueError(f"Durum '{current_state}' için geçerli bir geçiş bulunamadı.")
    
    print(f"Son Durum: {current_state}")
    return current_state

def evaluate_performance(results_df):
    """
    Karmaşıklık matrisi ve performans metriklerini hesaplar.
    """
    # Performans metriklerini hesaplamak için başlangıç değerleri
    DP = YN = YP = DN = 0

    # Gerçek ve tahmin edilen sınıfları kontrol et
    for i, row in results_df.iterrows():
        tahmin = row['FSM Tahmini'].lower()
        gercek = row['Gerçek Sınıf'].lower()
        
        if tahmin == "pozitif" and gercek == "pozitif":
            DP += 1
        elif tahmin == "pozitif" and gercek == "negatif":
            YP += 1
        elif tahmin == "negatif" and gercek == "pozitif":
            YN += 1
        elif tahmin == "negatif" and gercek == "negatif":
            DN += 1

    # Toplam tahmin sayısı
    toplam = DP + YP + YN + DN

    # Performans metrikleri
    dogruluk = (DP + DN) / toplam if toplam > 0 else 0
    kesinlik = DP / (DP + YP) if (DP + YP) > 0 else 0
    anma = DP / (DP + YN) if (DP + YN) > 0 else 0
    f1_olcutu = (2 * kesinlik * anma) / (kesinlik + anma) if (kesinlik + anma) > 0 else 0

    # Sonuçları yazdır
    print("Performans Değerleri:")
    print(f"Doğru Pozitif (DP): {DP}")
    print(f"Yanlış Pozitif (YP): {YP}")
    print(f"Yanlış Negatif (YN): {YN}")
    print(f"Doğru Negatif (DN): {DN}")
    print(f"Doğruluk: {dogruluk:.2f}")
    print(f"Kesinlik: {kesinlik:.2f}")
    print(f"Anma: {anma:.2f}")
    print(f"F1-Ölçütü: {f1_olcutu:.2f}")

    return {
        'Doğruluk': dogruluk,
        'Kesinlik': kesinlik,
        'Anma': anma,
        'F1-Ölçütü': f1_olcutu
    }

def highlight_cells_by_value(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active 

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            cell_value = str(cell.value).strip()
            if cell_value == "False":
                cell.fill = red_fill  
            elif cell_value == "True":
                cell.fill = green_fill  

    workbook.save(file_path)

def main(file_path):
    """
    Main fonksiyonu, dosyayı yükler, analizörleri başlatır, metni işler ve sonucu yazdırır.
    """
    # Excel dosyasını yükle
    df = pd.read_excel(file_path, engine='openpyxl')
    sentences = df['Cümle'].tolist()  # Cümle sütununu listeye al
    true_classes = df['Sınıf'].tolist()  # Sınıf sütununu listeye al

    # Analizörleri başlat
    analyzer = initialize_analyzers()

    # Sonuçları saklamak için bir liste
    results = []

    # Her cümle için işlem yap
    for i, sentence in enumerate(sentences):
        print("-" * 20)
        print(f"İşlenecek cümle {i + 1}: {sentence}")
        print("-" * 20)

        # Metni işle
        processed_text = process_text_with_zeyrek(sentence, analyzer)

        # İşlenmiş metni yazdır
        display_processed_text(processed_text)

        # FSM tanımı
        fsm = {
            "start": 
            {
                "ironic_punctuation": "Negatif",
                "no_ironic_punctuation": "positive_degil_control",
            },
            "positive_degil_control": 
            {
                "positive_degil": "Negatif",
                "no_positive_degil": "negative_degil_control",
            },
            "negative_degil_control": 
            {
                "negative_degil": "Pozitif",
                "no_positive_degil": "double_negation_control",
            },
            "double_negation_control": 
            {
                "double_negation": "Pozitif",
                "no_double_negation": "check_negative_beforeComma",
            },
            "check_negative_beforeComma": 
            {
                "negative_beforeComma": "Negatif",
                "no_negative_beforeComma": "check_positive_beforeComma",
            },
            "check_positive_beforeComma": 
            {
                "positive_beforeComma": "Pozitif",
                "no_positive_beforeComma": "check_negative_context",
            },
            "check_positive_context": 
            {
                "positive_words": "Pozitif",
                "no_positive_words": "check_negative_context",
            },
            "check_negative_context": 
            {
                "negative_words": "Negatif",
                "no_negative_words": "resolve_tie",
            },
            "resolve_tie": 
            {
                "positive_score_gte_negative": "Pozitif",
                "negative_score_gt_positive": "Negatif",
            },
        } 

        # Girdi verilerini analiz et
        input_data = analyze_sentence(processed_text)

        # FSM çalıştır
        predicted_class = run_fsm(fsm, input_data)
        print(f"FSM Tahmini: {predicted_class}")

        # Gerçek sınıfı al
        true_class = true_classes[i]
        print(f"Gerçek Sınıf: {true_class}")

        # Sonuçları sakla
        results.append({
            'Cümle': sentence,
            'FSM Tahmini': predicted_class,
            'Gerçek Sınıf': true_class,
            'Result?': predicted_class.lower() == true_class.lower()
        })
        print("-" * 30)

    # Sonuçları DataFrame olarak oluştur ve yazdır
    results_df = pd.DataFrame(results)
    print(results_df)

    # Sonuçları Excel dosyasına kaydet
    results_df.to_excel('tahmin_sonuclari.xlsx', index=False, engine='openpyxl')
    highlight_cells_by_value('tahmin_sonuclari.xlsx')
    print("Sonuçlar 'tahmin_sonuclari.xlsx' dosyasına kaydedildi.")

    # `Result?` değeri `False` olan satırları ayrı bir dosyaya kaydet
    errors_df = results_df[results_df['Result?'] == False]
    errors_df.to_excel('hatalar.xlsx', index=False, engine='openpyxl')
    print("Hatalar 'hatalar.xlsx' dosyasına kaydedildi.")


    # Performans değerlendirme
    performance_metrics = evaluate_performance(results_df)

    # Performans metriklerini yazdır
    print("\nPerformans Değerleri:")
    for metric, value in performance_metrics.items():
        print(f"{metric}: {value:.2f}")


# Ana fonksiyonu çağır
if __name__ == '__main__':
    file_path = 'ornek.xlsx' 
    main(file_path)

