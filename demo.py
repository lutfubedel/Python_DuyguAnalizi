import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
import zeyrek
import string
import pandas as pd
from nltk.corpus import stopwords
import openpyxl
from openpyxl.styles import PatternFill

# Gerekli NLTK paketlerini indirin
#nltk.download("all")
#nltk.download('punkt')
#nltk.download('stopwords')


def load_excel_file(file_path):
    """
    Excel dosyasını yükleyip, 'Cümle' sütununu döndüren fonksiyon.
    """
    df = pd.read_excel(file_path, engine='openpyxl')
    return df['Cümle'].tolist()  # 'Cümle' sütununu listeye dönüştür

def initialize_analyzers():
    """
    Zeyrek ve Stanza analizörlerini başlatan fonksiyon.
    """
    analyzer = zeyrek.MorphAnalyzer()
    return analyzer

def remove_punctuation(text):
    """
    Noktalama işaretlerini kaldıran fonksiyon.
    """
    words = word_tokenize(text)
    words = [word for word in words if word not in string.punctuation]
    return ' '.join(words)

def process_text_with_zeyrek(text, analyzer):
    """
    Zeyrek analizörü ile metni işleyen fonksiyon.
    """
    sentences = sent_tokenize(text)  # Metni cümlelere ayır
    processed_sentences = []

    # Türkçe StopWords listesini al
    turkish_stopwords = set(stopwords.words('turkish'))

    for sentence in sentences:
        clean_sentence = remove_punctuation(sentence)  # Noktalama işaretlerini kaldır
        words = word_tokenize(clean_sentence)  # Kelimelere ayır

        # Zeyrek ile kelimeleri köklere ve eklerine ayır
        sentence_info = []
        for word in words:
            word_lower = word.lower()  # Büyük-küçük harf duyarlılığını kaldır
            if word_lower in turkish_stopwords:  # Stopwords listesinde olan kelimeleri atla
                continue
            analysis = analyzer.analyze(word)
            if analysis:
                root = analysis[0][0].lemma  # Kelimenin kökü
                suffixes = "+".join(analysis[0][0].morphemes[1:])  # Kelimenin ekleri
                sentence_info.append({'Kelime': word, 'Kök': root, 'Ekler': suffixes if suffixes else 'Yok'})
            else:
                sentence_info.append({'Kelime': word, 'Kök': word, 'Ekler': 'Yok'})  # Eğer kök bulunamazsa olduğu gibi ekle

        processed_sentences.append(sentence_info)

    return processed_sentences

def negation_suffix_control(processed_text):
    # Olumsuzluk eklerini kontrol et
    for sentence in processed_text:
        for word_info in sentence:
            ekler = word_info["Ekler"].split("+")
            if word_info["Kök"] != "değil":
                if "Neg" in ekler or "Unable" in ekler or "Without" in ekler:
                    print(f"Olumsuzluk eki bulundu: {word_info['Kelime']}")
                    return True   
    return False

def double_negation_control(processed_text):
    # Eğer ilk olumsuzluk ekini bulmuşsak, "değil" kelimesini kontrol et
    for sentence in processed_text:
            for word_info in sentence:
                if word_info["Kök"] == "değil":
                    print("Değil kelimesi bulundu")
                    return True
        
    return False

def positive_polarity_control(processed_text, polarity_file):
    with open(polarity_file, "r", encoding="utf-8") as file:
        text_words = set(file.read().strip().lower().splitlines()) 

    for sentence in processed_text:
        for word_info in sentence:
            word = word_info["Kök"].lower()
            if word in text_words:
                ekler = word_info["Ekler"].split("+")
                if "Neg" or "Unable" in ekler or "Without" in ekler:
                    print(f"Olumsuz Kelime : {word_info['Kelime']}")
                    return False
                print(word)
                return True 
    return False

def negative_polarity_control(processed_text, polarity_file):
    with open(polarity_file, "r", encoding="utf-8") as file:
        text_words = set(file.read().strip().lower().splitlines()) 

    for sentence in processed_text:
        for word_info in sentence:
            word = word_info["Kök"].lower()
            ekler = word_info["Ekler"].split("+")
            if word in text_words or "Neg" in ekler or "Unable" in ekler or "Without" in ekler:
                print(f"Negative Word : {word_info['Kelime']}")
                return True 
    return False  

def positive_score_calculate(processed_text, polarity_file):
    score = 0
    if(positive_polarity_control(processed_text, polarity_file)):
        with open(polarity_file, "r", encoding="utf-8") as file:
            text_words = set(file.read().strip().lower().splitlines()) 

        for sentence in processed_text:
            for word_info in sentence:
                word = word_info["Kök"].lower()
                if word in text_words:
                    ekler = word_info["Ekler"].split("+")
                    if "Neg" in ekler:
                        print(f"Olumsuzluk eki bulundu: {word_info['Kelime']}")
                    else:
                        score +=1

                    print(word)

    return score

def negative_score_calculate(processed_text, polarity_file):
    score = 0
    if(negative_polarity_control(processed_text, polarity_file)):
        with open(polarity_file, "r", encoding="utf-8") as file:
            text_words = set(file.read().strip().lower().splitlines()) 

        for sentence in processed_text:
            for word_info in sentence:
                word = word_info["Kök"].lower()
                ekler = word_info["Ekler"].split("+")
                if word in text_words or "Neg" in ekler or "Unable" in ekler or "Without" in ekler:
                    score +=1
                    print(word)

    return score

def display_processed_text(processed_text):
    """
    İşlenmiş metni formatlı şekilde yazdıran fonksiyon.
    """
    print("----------------------------------------------------------------------------------------------------------")
    for i, sentence in enumerate(processed_text):
        print(f"Cümle {i+1}:")
        for word_info in sentence:
            print(f"Kelime: {word_info['Kelime']}, \nKök: {word_info['Kök']}, \nEkler: {word_info['Ekler']}\n--------------------------------------")
        print("----------------------------------------------------------------------------------------------------------")

def analyze_sentence(processed_text):
    # Flagleri başlat
    input_data = {
        "negation_suffix": False,
        "double_negation": False,
        "positive_words": False,
        "negative_words": False,
        "positive_score_gte_negative": False,
        "negative_score_gt_positive": False,

        "no_negation_suffix": True,
        "no_double_negation": True,
        "no_positive_words": True,
        "no_negative_words":True,
    }

    input_data["negation_suffix"] = negation_suffix_control(processed_text)
    input_data["double_negation"] = double_negation_control(processed_text)

    positive_polarity = positive_polarity_control(processed_text, "polarity_positive.txt")
    negative_polarity = negative_polarity_control(processed_text, "polarity_negative.txt")

    input_data["positive_words"] = positive_polarity
    input_data["negative_words"] = negative_polarity

    positive_score = positive_score_calculate(processed_text, "polarity_positive.txt")
    negative_score = negative_score_calculate(processed_text, "polarity_negative.txt")

    print("Positive Polarity : " , positive_polarity , "\t Score : ", positive_score)
    print("Negative Polarity : " , negative_polarity , "\t Score : ", negative_score)

    if(positive_score >= negative_score):
        input_data["positive_score_gte_negative"] = True
    
    if(negative_score > positive_score):
        input_data["negative_score_gt_positive"] = True

    return input_data

def run_fsm(fsm, input_data):
    """
    FSM'i çalıştırır ve bir son duruma ("positive" veya "negative") ulaşana kadar devam eder.
    """
    current_state = "start"

    while current_state not in ["Pozitif", "Negatif"]:
        # Mevcut durumdaki geçişleri al
        transitions = fsm.get(current_state, {})
        print(current_state)
        # Geçiş koşullarını kontrol et ve uygun bir sonraki duruma ilerle
        transition_found = False
        for condition, next_state in transitions.items():
            if condition in input_data and input_data[condition]:
                current_state = next_state
                transition_found = True
                break
        
        if not transition_found:
            raise ValueError(f"Durum '{current_state}' için geçerli bir geçiş bulunamadı.")
    
    return current_state

def evaluate_performance(results_df):
    """
    Karmaşıklık matrisi ve performans metriklerini hesaplar.
    """
    # Performans metriklerini hesaplamak için başlangıç değerleri
    DP = YN = YP = DN = 0

    # Gerçek ve tahmin edilen sınıfları kontrol et
    for i, row in results_df.iterrows():
        tahmin = row['FSM Tahmini'].lower()
        gercek = row['Gerçek Sınıf'].lower()
        
        if tahmin == "pozitif" and gercek == "pozitif":
            DP += 1
        elif tahmin == "pozitif" and gercek == "negatif":
            YP += 1
        elif tahmin == "negatif" and gercek == "pozitif":
            YN += 1
        elif tahmin == "negatif" and gercek == "negatif":
            DN += 1

    # Toplam tahmin sayısı
    toplam = DP + YP + YN + DN

    # Performans metrikleri
    dogruluk = (DP + DN) / toplam if toplam > 0 else 0
    kesinlik = DP / (DP + YP) if (DP + YP) > 0 else 0
    anma = DP / (DP + YN) if (DP + YN) > 0 else 0
    f1_olcutu = (2 * kesinlik * anma) / (kesinlik + anma) if (kesinlik + anma) > 0 else 0

    # Sonuçları yazdır
    print("Performans Değerleri:")
    print(f"Doğru Pozitif (DP): {DP}")
    print(f"Yanlış Pozitif (YP): {YP}")
    print(f"Yanlış Negatif (YN): {YN}")
    print(f"Doğru Negatif (DN): {DN}")
    print(f"Doğruluk: {dogruluk:.2f}")
    print(f"Kesinlik: {kesinlik:.2f}")
    print(f"Anma: {anma:.2f}")
    print(f"F1-Ölçütü: {f1_olcutu:.2f}")

    return {
        'Doğruluk': dogruluk,
        'Kesinlik': kesinlik,
        'Anma': anma,
        'F1-Ölçütü': f1_olcutu
    }

def highlight_cells_by_value(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active 

    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    for row in sheet.iter_rows():
        for cell in row:
            cell_value = str(cell.value).strip()
            if cell_value == "False":
                cell.fill = red_fill  
            elif cell_value == "True":
                cell.fill = green_fill  

    workbook.save(file_path)


def main(file_path):
    """
    Main fonksiyonu, dosyayı yükler, analizörleri başlatır, metni işler ve sonucu yazdırır.
    """
    # Excel dosyasını yükle
    df = pd.read_excel(file_path, engine='openpyxl')
    sentences = df['Cümle'].tolist()  # Cümle sütununu listeye al
    true_classes = df['Sınıf'].tolist()  # Sınıf sütununu listeye al

    # Analizörleri başlat
    analyzer = initialize_analyzers()

    # Sonuçları saklamak için bir liste
    results = []

    # Her cümle için işlem yap
    for i, sentence in enumerate(sentences):
        print("-" * 20)
        print(f"İşlenecek cümle {i + 1}: {sentence}")
        print("-" * 20)

        # Metni işle
        processed_text = process_text_with_zeyrek(sentence, analyzer)

        # İşlenmiş metni yazdır
        display_processed_text(processed_text)

        # FSM tanımı
        fsm = {
            "start": {
                "negation_suffix": "check_double_negation",
                "no_negation_suffix": "check_positive_context",
            },
            "check_double_negation": {
                "double_negation": "Pozitif",
                "no_double_negation": "check_negative_context",
            },
            "check_positive_context": {
                "positive_words": "Pozitif",
                "no_positive_words": "check_negative_context",
            },
            "check_negative_context": {
                "negative_words": "Negatif",
                "no_negative_words": "resolve_tie",
            },
            "resolve_tie": {
                "positive_score_gte_negative": "Pozitif",
                "negative_score_gt_positive": "Negatif",
            },
        }

        # Girdi verilerini analiz et
        input_data = analyze_sentence(processed_text)

        # FSM çalıştır
        predicted_class = run_fsm(fsm, input_data)
        print(f"FSM Tahmini: {predicted_class}")

        # Gerçek sınıfı al
        true_class = true_classes[i]
        print(f"Gerçek Sınıf: {true_class}")

        # Sonuçları sakla
        results.append({
            'Cümle': sentence,
            'FSM Tahmini': predicted_class,
            'Gerçek Sınıf': true_class,
            'Result?': predicted_class.lower() == true_class.lower()
        })
        print("-" * 30)

    # Sonuçları DataFrame olarak oluştur ve yazdır
    results_df = pd.DataFrame(results)
    print(results_df)

    # Sonuçları Excel dosyasına kaydet
    results_df.to_excel('tahmin_sonuclari.xlsx', index=False, engine='openpyxl')
    highlight_cells_by_value('tahmin_sonuclari.xlsx')
    print("Sonuçlar 'tahmin_sonuclari.xlsx' dosyasına kaydedildi.")

    # Performans değerlendirme
    performance_metrics = evaluate_performance(results_df)

    # Performans metriklerini yazdır
    print("\nPerformans Değerleri:")
    for metric, value in performance_metrics.items():
        print(f"{metric}: {value:.2f}")


# Ana fonksiyonu çağır
if __name__ == '__main__':
    file_path = 'ornek.xlsx' 
    main(file_path)

